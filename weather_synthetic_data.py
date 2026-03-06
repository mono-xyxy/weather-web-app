import os
import random
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
ROWS_PER_TYPE = 200_000  # 200k x 5 = 1,000,000 total
OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "weather.xlsx")

PALETTE = {
    "Sunny":        {"header_bg": "FFD700", "header_fg": "7B4F00"},
    "Cloudy":       {"header_bg": "B0BEC5", "header_fg": "1A1A1A"},
    "Rain":         {"header_bg": "1565C0", "header_fg": "FFFFFF"},
    "Thunderstorm": {"header_bg": "4A148C", "header_fg": "FFFFFF"},
    "Fog":          {"header_bg": "78909C", "header_fg": "FFFFFF"},
}

UNITS = {
    "temperature":   "°C",
    "humidity":      "%",
    "pressure":      "hPa",
    "wind_speed":    "km/h",
    "precipitation": "mm",
    "cloud_cover":   "%",
    "visibility":    "km",
    "weather":       "",
}

# ── Generators ────────────────────────────────────────────────────────────────
def make_sunny():
    return {
        "temperature":   round(random.uniform(25, 35),    2),
        "humidity":      round(random.uniform(30, 60),    2),
        "pressure":      round(random.uniform(1010, 1025),2),
        "wind_speed":    round(random.uniform(3, 15),     2),
        "precipitation": 0.0,
        "cloud_cover":   round(random.uniform(0, 20),     2),
        "visibility":    round(random.uniform(8, 15),     2),
        "weather": "Sunny",
    }

def make_cloudy():
    return {
        "temperature":   round(random.uniform(15, 25),    2),
        "humidity":      round(random.uniform(50, 75),    2),
        "pressure":      round(random.uniform(1000, 1015),2),
        "wind_speed":    round(random.uniform(5, 20),     2),
        "precipitation": round(random.uniform(0, 2),      2),
        "cloud_cover":   round(random.uniform(50, 85),    2),
        "visibility":    round(random.uniform(5, 10),     2),
        "weather": "Cloudy",
    }

def make_rain():
    return {
        "temperature":   round(random.uniform(10, 22),   2),
        "humidity":      round(random.uniform(70, 95),   2),
        "pressure":      round(random.uniform(990, 1010),2),
        "wind_speed":    round(random.uniform(10, 25),   2),
        "precipitation": round(random.uniform(1, 20),    2),
        "cloud_cover":   round(random.uniform(80, 100),  2),
        "visibility":    round(random.uniform(2, 6),     2),
        "weather": "Rain",
    }

def make_thunderstorm():
    return {
        "temperature":   round(random.uniform(10, 25),   2),
        "humidity":      round(random.uniform(80, 100),  2),
        "pressure":      round(random.uniform(975, 995), 2),
        "wind_speed":    round(random.uniform(30, 80),   2),
        "precipitation": round(random.uniform(15, 60),   2),
        "cloud_cover":   round(random.uniform(90, 100),  2),
        "visibility":    round(random.uniform(0.5, 3),   2),
        "weather": "Thunderstorm",
    }

def make_fog():
    return {
        "temperature":   round(random.uniform(5, 20),     2),
        "humidity":      round(random.uniform(85, 100),   2),
        "pressure":      round(random.uniform(1005, 1020),2),
        "wind_speed":    round(random.uniform(0, 8),      2),
        "precipitation": round(random.uniform(0, 0.5),    2),
        "cloud_cover":   round(random.uniform(70, 95),    2),
        "visibility":    round(random.uniform(0.05, 1),   2),
        "weather": "Fog",
    }

GENERATORS = {
    "Sunny":        make_sunny,
    "Cloudy":       make_cloudy,
    "Rain":         make_rain,
    "Thunderstorm": make_thunderstorm,
    "Fog":          make_fog,
}

# ── Generate data ─────────────────────────────────────────────────────────────
print(f"Generating {ROWS_PER_TYPE:,} rows x {len(GENERATORS)} types = {ROWS_PER_TYPE * len(GENERATORS):,} total rows...")

weather_dfs = {}
for wtype, fn in GENERATORS.items():
    print(f"  • {wtype}...", end=" ", flush=True)
    weather_dfs[wtype] = pd.DataFrame([fn() for _ in range(ROWS_PER_TYPE)])
    print(f"{len(weather_dfs[wtype]):,} rows done")

all_df = pd.concat(weather_dfs.values(), ignore_index=True)
print(f"Total: {len(all_df):,} rows\n")

# ── Write all data to Excel (data only, fast) ─────────────────────────────────
print(f"Writing data to Excel: {OUTPUT}")
print("  (This may take 1-2 minutes for 1M rows...)")

with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
    all_df.to_excel(writer, sheet_name="All Data", index=False)
    print("  v All Data sheet written")
    for wtype, df in weather_dfs.items():
        df.to_excel(writer, sheet_name=wtype, index=False)
        print(f"  v {wtype} sheet written")

# ── Style HEADERS ONLY (fast) ─────────────────────────────────────────────────
print("\nStyling headers...")
wb = load_workbook(OUTPUT)
COLUMNS = list(all_df.columns)

def style_header(ws, wtype):
    colors   = PALETTE.get(wtype, {"header_bg": "37474F", "header_fg": "FFFFFF"})
    hdr_fill = PatternFill("solid", fgColor=colors["header_bg"])
    hdr_font = Font(name="Arial", bold=True, color=colors["header_fg"], size=10)
    center   = Alignment(horizontal="center", vertical="center")
    border   = Border(bottom=Side(style="medium", color=colors["header_bg"]))

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell       = ws.cell(row=1, column=col_idx)
        unit       = UNITS.get(col_name, "")
        cell.value = f"{col_name} ({unit})" if unit else col_name
        cell.fill  = hdr_fill
        cell.font  = hdr_font
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(cell.value) + 4, 14)

    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = colors["header_bg"]
    ws.row_dimensions[1].height = 20

for sheet_name in wb.sheetnames:
    wtype = sheet_name if sheet_name in PALETTE else "All Data"
    style_header(wb[sheet_name], wtype)
    print(f"  v {sheet_name} header styled")

wb["All Data"].sheet_properties.tabColor = "37474F"
wb.save(OUTPUT)

print(f"\nDone!")
print(f"   File  : {OUTPUT}")
print(f"   Sheets: {', '.join(wb.sheetnames)}")
print(f"   Rows  : {len(all_df):,}")